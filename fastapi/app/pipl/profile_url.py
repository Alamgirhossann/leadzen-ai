import asyncio
import hashlib
import urllib
from typing import Optional, List, Dict, Union
from urllib.parse import urlparse, urlunparse, urlencode
from fastapi import HTTPException
from starlette import status

from loguru import logger
from pydantic import BaseModel, HttpUrl

from app.credits.email import get_email_by_hash_key
from app.history import add_history
from app.config import API_CONFIG_PIPL_BASE_URL, API_CONFIG_PIPL_API_KEY, API_CONFIG_EXCEL_FILE_PATH
from app.credits.admin import deduct_credit
from app.pipl.common import write_to_file, search_all, search_all_by_pipl, search_all_by_texAu
from datetime import date
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.profile_search import get_profile_search, add_profile
from app.users import get_user
from app.users import User


class PiplDetailsFromProfileUrlRequest(BaseModel):
    profile_urls: List[str]
    filename: Optional[str] = None
    hash_key_list: Optional[List[Dict]] = None
    user: User
    export_type: Optional[str] = None
    search_id: Optional[str] = None
    search_index: Optional[List[Dict]] = None


class PiplDetailsFromProfileUrlResponse(BaseModel):
    filename: str


async def execute_task(request: PiplDetailsFromProfileUrlRequest):
    responses = None
    logger.debug(f"{request.export_type=}")
    if request.export_type == "texAu":
        profile_urls = list(set(request.profile_urls))  # remove duplicates
        profile_urls = [x for x in profile_urls if x]  # remove empty profile_urls

        urls = [
            f"{API_CONFIG_PIPL_BASE_URL}/?{urlencode({'url': profile_url, 'key': API_CONFIG_PIPL_API_KEY, 'match_requirements': 'phones'})}"
            for profile_url in profile_urls
            if profile_url
        ]

        if not (
                responses := await search_all_by_texAu(
                    urls=urls,
                    slugs=profile_urls,
                    hash_key_list=request.hash_key_list,
                    search_id=request.search_id,
                    search_index=request.search_index,
                    user=request.user,
                )
        ):
            logger.error("Error Getting Data")
            return
        if not any(responses):
            logger.error("no valid responses found")
            return
        await write_to_file(responses=responses, filename=request.filename)

    elif request.export_type == "PIPL":
        logger.debug(f"In PIPL")
        if not (
                responses := await search_all_by_pipl(
                    urls=request.profile_urls,
                    slugs=request.profile_urls,
                    hash_key_list=request.hash_key_list,
                    search_id=request.search_id,
                    search_index=request.search_index,
                    user=request.user,
                )
        ):
            logger.error("Error Getting Data")
            return
        if not any(responses):
            logger.error("no valid responses found")
            return
        await write_to_file(responses=responses, filename=request.filename)
    else:
        logger.debug(f"In Bulk csv")
        await handle_bulk_search(request=request)


async def handle_bulk_search(request: PiplDetailsFromProfileUrlRequest):
    logger.debug(f"{len(request.profile_urls[0])=}>>>{request.profile_urls[0]}")
    profile_urls = list(set(request.profile_urls))  # remove duplicates
    profile_urls = [x for x in profile_urls if x]  # remove empty profile_urls

    urls = [
        f"{API_CONFIG_PIPL_BASE_URL}/?{urlencode({'url': profile_url, 'key': API_CONFIG_PIPL_API_KEY, 'match_requirements': 'phones'})}"
        for profile_url in profile_urls
        if profile_url
    ]

    def filter_results(results: List[Dict]) -> Dict:
        found_in_history = [x for x in results if x[0]]
        not_found_in_history = [x for x in results if not x[0]]
        return {"found": found_in_history, "not_found": not_found_in_history}

    history_search_result = await search_in_history(urls=urls, user=request.user)

    history_search_result_1 = filter_results(history_search_result)
    logger.debug(f"{history_search_result_1=}")

    found_urls = []
    founds = history_search_result_1.get("found")
    logger.debug(f"In found>@@@@@@@@@@@@@@{type(founds)=}>>>>{len(founds)=}>>>>")

    pipl_search_results_found = None
    filtered_list = []
    for items in founds:
        logger.debug(f"In founds>>>>>>>>>{type(items)=}>>>{items[1]=}")
        found_urls.append(items[1])
    # logger.debug(f"%%%%%%{len(found_urls}")
    if found_urls:
        if not (
                pipl_search_results_found := await search_all(urls=found_urls, slugs=found_urls)
        ):
            logger.error("Error Getting Data")
            return
        logger.debug(f"&&&&&&&&{pipl_search_results_found=}")

        for items in pipl_search_results_found:
            items = tuple(items)
            logger.debug(f"{type(items)=}>>>>")
            filtered_list.append(items[1])
            print(len(items))

    if not_founds := history_search_result_1.get("not_found"):

        logger.debug(f"{not_founds=}>>{type(not_founds)=}>>>{found_urls=}")
        not_found_urls = None
        if not found_urls and not not_founds and not any(not_founds):
            logger.debug(f"not any>>>>>1111")
            return
        # not_found_urls = [item[1] for item in not_founds]
        not_found_urls_only = []
        if not_founds:
            for item in not_founds:
                logger.debug(f"{item=}>>>")
                not_found_urls_only.append(item[1])
        not_found_urls = [x[1] for x in not_founds if x]
        logger.debug(f"{not_found_urls_only=}")
        pipl_search_results = None
        if not_found_urls_only:
            if not (
                    pipl_search_results := await search_all(urls=not_found_urls_only, slugs=not_found_urls_only)
            ):
                logger.error("Error Getting Data")
                return
        logger.debug(f"Before pipl_search_result>>>{len(pipl_search_results)=}>>>>")
        if not any(pipl_search_results):
            logger.error("no valid responses found")
            return
        pipl_search_results_res = []

        for items in pipl_search_results:
            logger.debug(f"{items[0]=}")
            pipl_search_results_res.append(items[0])
            filtered_list.append(items[1])

        await check_result_exists_in_profile_search_history(
                responses=pipl_search_results_res, urls=urls, user=request.user)
        # ):
        #     await check_result_exists_in_email_search_history(
        #         responses=filtered_list, urls=urls, user=request.user
        #     )

    logger.debug(f"{type(filtered_list)=}>>>>>>{filtered_list=}")

    return await write_to_file(responses=filtered_list, filename=request.filename)


async def search_in_history(urls: List[str], user: User) -> tuple[
    Union[HTTPException, BaseException], Union[HTTPException, BaseException], Union[HTTPException, BaseException],
    Union[HTTPException, BaseException], Union[
        HTTPException, BaseException]]:
    # coroutines_url_hash_key = [make_url_hash_key(url=x) for x in urls]

    async def check_hash_key_exists_in_profile_search_history(
            url_hash: str, user: User
    ) -> Optional[Dict]:
        # check for url hashes and respond with first result
        if not (response_profile_search := await get_profile_search(url_hash, user)):
            return None
        return response_profile_search

    async def check_one_result(url: str, user: User):
        parsed = urlparse(url)
        url_split_temp = parsed.query.split("&")[0]
        split_url = url_split_temp.split("=")[1]
        url_hash_key_initial = urllib.parse.unquote(split_url)

        url_hash_key = await make_url_hash_key(url=url_hash_key_initial)

        logger.debug(f">>>{ len(url_hash_key_initial)=}>>>{url_hash_key_initial=}>>>{url_hash_key=}")
        if not (
                result := await check_hash_key_exists_in_profile_search_history(
                    url_hash=url_hash_key, user=user
                )
        ):
            return None, url

        return result, url

    coroutines = [check_one_result(url=x, user=user) for x in urls]

    results = await asyncio.gather(*coroutines)
    logger.debug(f"{results=}")
    return results
    # return filter_results(results)


async def make_url_hash_key(url: str) -> str:
    logger.debug(f" In make_url_hash_key : {url=}")
    encoded = url.encode()
    result = hashlib.sha256(encoded)
    return result.hexdigest()


async def check_result_exists_in_profile_search_history(
        responses: List[Dict], urls: List[str], user: User
):
    # def make_url_hash_key(url: str) -> str:
    #     pass

    def make_result_hash_key(result: Dict) -> str:
        logger.debug(f"{str(result)}")
        encoded = str(result).encode()
        result = hashlib.sha256(encoded)
        return result.hexdigest()

    async def add_to_history(result: Dict, user):

        logger.debug(f"{result=}")
        return await add_history('Bulk', 'Bulk', [result], user)

    async def add_to_profile_credit(search_id, result, user):
        logger.debug(f"{search_id=}>>>{type(result)=}>>>{result}")
        result = {
            k: v for list_item in result for (k, v) in list_item.items()
        }
        phones = result.get("phones")
        phone_list = []
        for phone in phones:
            phone_list.append(str(phone.get("number")))
        # logger.debug(f"{phone_list=}>>>{search_index}>>>{type(search_index)}")
        # bulk_add_credit_profile(search_id=request.search_id, phone_numbers=request.phone_numbers,
        #                         search_index=request.search_index, user=user)

    async def check_hash_key_exists_in_profile_search_history(
            result_hash: str
    ) -> Optional[Dict]:
        if not (response_profile_search := await get_profile_search(result_hash, user)):
            return None
        logger.debug(f"{type(response_profile_search)}>>>>{response_profile_search=}")
        return response_profile_search

    async def add_to_profile_search(search_type: str, hash_key: str, search_results: Dict, user):
        request_add_profile = {
            "search_type": search_type,
            "hash_key": hash_key,
            "search_results": [search_results],
        }
        return await add_profile(request_add_profile, user)

    async def check_one_result(url: str, result: Dict):
        logger.debug(f"{url=}")
        parsed = urlparse(url)
        url_split_temp = parsed.query.split("&")[0]
        split_url = url_split_temp.split("=")[1]

        url_hash_key = await make_url_hash_key(url=urllib.parse.unquote(split_url))
        # logger.debug(f"in after result {result=}")
        result_hash_key = make_result_hash_key(result=result)
        logger.debug(f"{result_hash_key=}")
        if (exists_response := await check_hash_key_exists_in_profile_search_history(
                result_hash=result_hash_key
        )):
            logger.debug("profile found by serach result hash key>>")
            return None
        logger.debug("profile not found>>>")
        logger.debug(f"{exists_response=}")
        await add_to_profile_search(
            search_type="texAu",
            hash_key=url_hash_key,
            search_results=result, user=user
        )
        await deduct_user_profile_credit("PROFILE", user)

        exists_response = result
        logger.debug(f"{exists_response=}")
        return exists_response

    coroutines = [check_one_result(url=x[0], result=x[1]) for x in zip(urls, responses)]

    results = await asyncio.gather(*coroutines)
    logger.debug(f"{results=}")
    return results


async def deduct_user_profile_credit(credit_type, user):
    user_response = await get_user(user)
    if credit_type == "PROFILE":
        if user_response and user_response.profile_credit <= 0:
            logger.warning("Insufficient Credits")
            raise HTTPException(
                status_code=status.HTTP_402_PAYMENT_REQUIRED,
                detail="Insufficient Credits",
            )
    if credit_type == "EMAIL":
        if user_response and user_response.email_credit <= 0:
            logger.warning("Insufficient Credits")
            raise HTTPException(
                status_code=status.HTTP_402_PAYMENT_REQUIRED,
                detail="Insufficient Credits",
            )
    return await deduct_credit(credit_type, user_response)


def add_excel_template_to_file(outgoing_filename, user):
    if (outgoing_filename == ""):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str("Excel file not found"),
        )
    else:
        wb_orginal = load_workbook(f'{outgoing_filename}')
        if (wb_orginal == ""):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=str("Error in Work book Loading"),
            )
        else:
            try:
                wb_template = load_workbook(f'{API_CONFIG_EXCEL_FILE_PATH}/LeadZen_template.xlsx')
                index_sheet = wb_template.get_sheet_by_name('Index')
                data = wb_orginal.get_sheet_by_name('Sheet1')
                leads_sheet = wb_template.get_sheet_by_name('Leads')
                data_sheet_row_count = data.max_row
                data_sheet_col_count = data.max_column
                for row in range(1, data_sheet_row_count + 1):
                    for col in range(1, data_sheet_col_count + 1):
                        c = data.cell(row=row, column=col)
                        leads_sheet.cell(row=row + 5, column=col).value = c.value
                for rows in leads_sheet.iter_rows(min_row=6, max_row=6, min_col=1):
                    for cell in rows:
                        cell.fill = PatternFill(fgColor="B4C9D9", patternType="solid")
                no_of_record = index_sheet['J8']
                requested_by = index_sheet['J10']
                request_date = index_sheet['J12']
                no_of_record.value = data_sheet_row_count - 1
                requested_by.value = user.username
                today = date.today()
                request_date.value = today.strftime("%d/%m/%Y")
                wb_template.save(outgoing_filename)
                return outgoing_filename
            except Exception as e:
                logger.critical("Error>>>" + str(e))


async def check_result_exists_in_email_search_history(responses, urls, user):
    logger.debug(f"check_result_exists_in_email_search_history{type(responses)=}>>>>{responses=}>>>>{urls=}>>>>{user}")

    async def check_hash_key_exists_in_email_search_history(url_hash, user):
        if url_hash and user:
            email_response = await get_email_by_hash_key(url_hash, user)

            if not email_response:
                await deduct_credit("EMAIL", user)

    async def check_one_result(url, result):
        logger.debug(f"{url=}>>>{result=}")
        parsed = urlparse(url)
        url_split_temp = parsed.query.split("&")[0]
        split_url = url_split_temp.split("=")[1]
        url_hash_key_initial = urllib.parse.unquote(split_url)

        url_hash_key = await make_url_hash_key(url=url_hash_key_initial)

        logger.debug(f">>>{ len(url_hash_key_initial)=}>>>{url_hash_key_initial=}>>>{url_hash_key=}")
        if not (
                result := await check_hash_key_exists_in_email_search_history(
                    url_hash=url_hash_key, user=user
                )
        ):
            return None, url

        return result, url

    coroutines = [check_one_result(url=x[0], result=x[1]) for x in zip(urls, responses)]

    results = await asyncio.gather(*coroutines)
    logger.debug(f"{results=}")
    return results
