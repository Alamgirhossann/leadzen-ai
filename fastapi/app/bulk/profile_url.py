from typing import List

import pandas as pd
from app.bulk.common import BulkRequest, wait_and_check_for_filename
from app.pipl.profile_url import (
    execute_task as execute_profile_task,
    PiplDetailsFromProfileUrlRequest,
)
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class BulkProfileUrlRequest(BulkRequest):
    urls: List[str]


async def handle_bulk_profile_urls(request: BulkProfileUrlRequest):
    await execute_profile_task(
        request=PiplDetailsFromProfileUrlRequest(
            profile_urls=request.urls, filename=request.outgoing_filename
        )
    )
    filename=request.outgoing_filename
    wb = load_workbook(filename)
    wb.create_sheet('Index', 0)
    index_sheet = wb.get_sheet_by_name('Index')
    index_sheet.merge_cells('A1:W36')
    img1 = openpyxl.drawing.image.Image(f'./app/Excel/Index.png')
    img1.anchor = 'A1'
    index_sheet.add_image(img1)
    wb.create_sheet('Disclaimer',2)
    dis_sheet = wb.get_sheet_by_name('Disclaimer')
    dis_sheet.merge_cells('A1:X31')
    img2 = openpyxl.drawing.image.Image(f'./app/Excel/Disclaimer.png')
    img2.anchor = 'A1'
    dis_sheet.add_image(img2)
    financial_advisor = wb.get_sheet_by_name('Sheet1')
    financial_advisor.title = "Financial Advisor"
    for rows in financial_advisor.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.fill = PatternFill(fgColor="B4C9D9", patternType="solid")
    # number_of_rows = financial_advisor.max_row
    number_of_col = financial_advisor.max_column
    last_column = chr(65 + number_of_col - 1)
    lastcell = last_column + str(number_of_col)
    financial_advisor.move_range(f"A1:{lastcell}", rows=8, cols=0)
    financial_advisor.merge_cells('A1:W8')
    img3 = openpyxl.drawing.image.Image(f'./app/Excel/FA_img.png')
    img3.anchor = 'A1'
    financial_advisor.add_image(img3)
    wb.save(filename)

    await wait_and_check_for_filename(
        request=BulkRequest(
            incoming_filename=request.incoming_filename,
            outgoing_filename=request.outgoing_filename,
            user=request.user,
        )
    )
